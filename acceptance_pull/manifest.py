from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from acceptance_pull.models import ManifestAsset


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_manifest(path: Path) -> dict[str, ManifestAsset]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(values_only=True))
        headers = {str(value).strip(): index for index, value in enumerate(header_row) if value is not None}
        missing = {"asset_id", "scene", "task"} - set(headers)
        if missing:
            raise ValueError(f"manifest missing required columns: {', '.join(sorted(missing))}")

        assets: dict[str, ManifestAsset] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            asset_id = normalize_id(row[headers["asset_id"]])
            if not asset_id:
                continue
            scene = str(row[headers["scene"]] or "").strip()
            task = str(row[headers["task"]] or "").strip()
            assets[asset_id] = ManifestAsset(asset_id=asset_id, scene=scene, task=task)
        return assets
    finally:
        wb.close()
